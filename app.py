# app.py
"""Streamlit application for balanced group distribution."""

import sys
import os
import io

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import streamlit as st
import pandas as pd
from gender_ai import detect_gender, is_name_recognized, get_namsor_api_key
from generator import generate_groups

# Page configuration
st.set_page_config(page_title="РАСХОДИМСЯ ПО ГРУППАМ!", layout="wide")
st.title("🔥 РАСХОДИМСЯ ПО ГРУППАМ!")

# Constants
DATA_KEY = "residents_data"
WIDGET_KEY = "residents_widget"
DEFAULT_COLUMNS = ["Имя", "Пол", "Роль", "🚦 Статус"]
ROLE_OPTIONS = ["Обычный", "ВПИ", "Новичок"]
GENDER_OPTIONS = ["M", "F"]


def initialize_session_state() -> None:
    """Initialize session state with default DataFrame if not present."""
    if DATA_KEY not in st.session_state or not isinstance(
        st.session_state.get(DATA_KEY), pd.DataFrame
    ):
        st.session_state[DATA_KEY] = pd.DataFrame(columns=DEFAULT_COLUMNS)
    if "namsor_debug_log" not in st.session_state:
        st.session_state["namsor_debug_log"] = []


def migrate_roles() -> bool:
    """Migrate old role names to new Russian names. Returns True if migration occurred."""
    df = st.session_state[DATA_KEY]
    if df.empty or "Роль" not in df.columns:
        return False
    
    role_mapping = {"regular": "Обычный", "expert": "ВПИ", "newbie": "Новичок"}
    if df["Роль"].isin(role_mapping.keys()).any():
        df["Роль"] = df["Роль"].replace(role_mapping)
        st.session_state[DATA_KEY] = df.copy()
        if WIDGET_KEY in st.session_state:
            del st.session_state[WIDGET_KEY]
        return True
    return False


def get_current_dataframe() -> pd.DataFrame:
    """Get current dataframe from widget or session state."""
    if WIDGET_KEY in st.session_state and isinstance(
        st.session_state[WIDGET_KEY], pd.DataFrame
    ):
        return st.session_state[WIDGET_KEY]
    return st.session_state.get(DATA_KEY, pd.DataFrame(columns=DEFAULT_COLUMNS))


def add_bulk_names(names_text: str) -> tuple[bool, str]:
    """Add multiple names from text input.
    
    Returns:
        Tuple of (success, message)
    """
    names = [n.strip() for n in names_text.splitlines() if n.strip()]
    if not names:
        return False, "Введите имена для добавления."
    
    df = get_current_dataframe()
    existing_names = set(df["Имя"].dropna().str.strip().tolist())
    unique_names = list(dict.fromkeys(names))
    to_add = [n for n in unique_names if n not in existing_names]
    
    if not to_add:
        return False, "Все имена уже есть в таблице."
    
    # Detect genders and statuses
    genders = []
    statuses = []
    debug_messages = []
    for name in to_add:
        gender_result, success, debug_msg, request_details = detect_gender(name)
        genders.append(gender_result)
        if success:
            statuses.append("Пол определён через ИИ")
        else:
            statuses.append("🔴 Не удалось определить пол")
        debug_messages.append((name, debug_msg, request_details))
        # Add to global debug log with full request/response details
        st.session_state["namsor_debug_log"].append({
            "name": name, 
            "message": debug_msg, 
            "success": success,
            "request_details": request_details
        })
    
    new_rows = pd.DataFrame({
        "Имя": to_add,
        "Пол": genders,
        "Роль": "Обычный",
        "🚦 Статус": statuses
    })
    
    st.session_state[DATA_KEY] = pd.concat([df, new_rows], ignore_index=True)
    if WIDGET_KEY in st.session_state:
        del st.session_state[WIDGET_KEY]
    
    return True, f"✅ Добавлено: {len(to_add)}"


def auto_detect_genders() -> None:
    """Auto-detect gender for all participants."""
    df = get_current_dataframe().copy()
    if df.empty:
        return
    
    # Process each row to get gender, success status, and debug message
    results = df["Имя"].astype(str).apply(detect_gender)
    df["Пол"] = results.apply(lambda x: x[0])
    
    # Collect debug messages
    debug_messages = []
    
    # Set status based on whether Namsor successfully detected the gender
    def get_status_and_debug(name):
        _, success, debug_msg, request_details = detect_gender(name)
        # Add to global debug log with full request/response details
        st.session_state["namsor_debug_log"].append({
            "name": name, 
            "message": debug_msg, 
            "success": success,
            "request_details": request_details
        })
        if success:
            return ("Пол определён через ИИ", debug_msg)
        else:
            return ("🔴 Не удалось определить пол", debug_msg)
    
    status_results = df["Имя"].apply(get_status_and_debug)
    df["🚦 Статус"] = status_results.apply(lambda x: x[0])
    debug_messages = status_results.apply(lambda x: x[1]).tolist()
    
    st.session_state[DATA_KEY] = df
    if WIDGET_KEY in st.session_state:
        del st.session_state[WIDGET_KEY]


def parse_limits(limits_text: str) -> list:
    """Parse limits from text input into LimitConstraint objects.
    
    Supports two formats:
    1. Many-to-many (default): "Name1, Name2, Name3" - all cannot be together
       Any subset of 2+ members from this group is forbidden.
       
    2. One-to-many: "Name1 -> Name2, Name3" - Name1 cannot be with Name2 or Name3,
       but Name2 and Name3 can be together.
    
    Each constraint should be on a separate line.
    """
    from generator import LimitConstraint
    
    constraints = []
    for line in limits_text.splitlines():
        line = line.strip()
        if not line:
            continue
        
        # Check for one-to-many format (with arrow)
        if '->' in line:
            parts = line.split('->', 1)
            source = parts[0].strip()
            targets_str = parts[1].strip()
            targets = [t.strip() for t in targets_str.split(',') if t.strip()]
            if source and targets:
                try:
                    constraint = LimitConstraint.create_one_to_many(source, targets)
                    constraints.append(constraint)
                except ValueError as e:
                    pass  # Skip invalid constraints
        else:
            # Many-to-many format (comma-separated list)
            members = [m.strip() for m in line.split(',') if m.strip()]
            if len(members) >= 2:
                try:
                    constraint = LimitConstraint.create_many_to_many(members)
                    constraints.append(constraint)
                except ValueError as e:
                    pass  # Skip invalid constraints
    
    return constraints


def render_resident_table() -> pd.DataFrame:
    """Render the resident data editor table."""
    st.subheader("📝 Резиденты")
    
    return st.data_editor(
        st.session_state[DATA_KEY].copy(),
        key=WIDGET_KEY,
        column_config={
            "Имя": st.column_config.TextColumn("Имя", required=True),
            "Пол": st.column_config.SelectboxColumn(
                "Пол", options=GENDER_OPTIONS, required=True, default="M"
            ),
            "Роль": st.column_config.SelectboxColumn(
                "Роль", options=ROLE_OPTIONS, required=True, default="Обычный"
            ),
            "🚦 Статус": st.column_config.TextColumn("Статус", width="small", disabled=True),
        },
        hide_index=True,
        width="stretch",
        num_rows="dynamic"
    )


def render_generation_settings() -> dict:
    """Render generation settings and return values."""
    with st.expander("⚙️ Настройки генерации"):
        return {
            "num_groups": st.number_input("Количество групп", min_value=1, value=2),
            "strict_roles": st.checkbox("Строгий баланс ролей", value=True),
            "strict_genders": st.checkbox("Строгий баланс полов", value=True),
            "seed": st.number_input("Seed (опционально)", value=None, step=1),
        }


def main():
    """Main application entry point."""
    # Initialize session state
    initialize_session_state()
    
    # Check for missing API key and show warning
    if not get_namsor_api_key():
        st.warning("⚠️ **AI-определение пола не работает**: отсутствует переменная окружения `NAMSOR_API_KEY`. "
                  "Пожалуйста, установите её для автоматического определения пола.")
    
    # Run role migration if needed
    if migrate_roles():
        st.rerun()
    
    # Bulk import section
    with st.expander("📋 Массовое добавление резидентов", expanded=True):
        bulk_text = st.text_area(
            "Вставьте список имён (каждое с новой строки)", height=80
        )
        if st.button("➕ Добавить в таблицу", width="stretch"):
            success, message = add_bulk_names(bulk_text)
            if success:
                st.success(message)
                st.rerun()
            else:
                st.warning(message)
    
    # Auto-detect gender button
    if st.button("🔍 Авто-определить пол у всех", type="secondary", width="stretch"):
        auto_detect_genders()
        st.rerun()
    
    # Warning for unrecognized names
    current_df = st.session_state[DATA_KEY]
    if "🚦 Статус" in current_df.columns:
        undetected = current_df[current_df["🚦 Статус"].str.contains("Не удалось определить пол", na=False)]
        if not undetected.empty:
            st.warning(f"🔴 Проверьте вручную: {', '.join(undetected['Имя'])}")
    
    # Expandable Namsor API details section (closed by default)
    with st.expander("📡 Детали коммуникации с Namsor API", expanded=False):
        if st.session_state["namsor_debug_log"]:
            # Display as plain text with all entries showing full HTTP request/response details
            for entry in st.session_state["namsor_debug_log"]:
                if entry["success"]:
                    status_text = "Пол определен ИИ"
                    status_color = "green"
                else:
                    status_text = "Пол не мог быть определён ИИ"
                    status_color = "red"
                
                # Format request and response details as plain text
                request_details = entry.get("request_details", {})
                request_info = request_details.get("request", {})
                response_info = request_details.get("response", {})
                
                # Build detailed info string
                details_lines = []
                details_lines.append(f"Имя: {entry['name']}")
                details_lines.append(f"Статус: {status_text}")
                details_lines.append("")
                
                if request_info:
                    details_lines.append("Request:")
                    details_lines.append(f"  URL: {request_info.get('url', 'N/A')}")
                    details_lines.append(f"  Method: {request_info.get('method', 'N/A')}")
                    details_lines.append(f"  Headers: {request_info.get('headers', {})}")
                    details_lines.append(f"  Payload: {request_info.get('payload', {})}")
                
                if response_info:
                    details_lines.append("")
                    details_lines.append("Response:")
                    if "error" in response_info:
                        details_lines.append(f"  Error: {response_info['error']}")
                    else:
                        details_lines.append(f"  Status Code: {response_info.get('status_code', 'N/A')}")
                        details_lines.append(f"  Body: {response_info.get('body', {})}")
                
                details_lines.append("")
                details_lines.append("-" * 50)
                details_lines.append("")
                
                st.text("\n".join(details_lines))
        else:
            st.info("Пока нет записей о коммуникации с Namsor API. Добавьте имена или используйте авто-определение пола.")
    
    # Resident table
    editor_df = render_resident_table()
    
    # Limits section
    st.subheader("🌍 Границы")
    st.markdown("""
    **Как работают ограничения:**  
    
    Есть два типа ограничений:
    
    1. **Многие-ко-многим** (по умолчанию): перечислите имена через запятую — никакие двое из этой группы не должны быть вместе.
       - Пример: `Олег С, Леша Ч, Петя И` → все трое должны быть в разных группах
    
    2. **Один-ко-многим**: используйте стрелку `->` чтобы указать, что один человек не должен быть с другими, но те могут быть вместе.
       - Пример: `Аня К -> Петя О, Маша И` → Аня не должна быть с Петей или Машей, но Петя и Маша могут быть вместе
    
    **Примеры:**
    - *Пара (многие-ко-многим):* `Олег С, Леша Ч` → Олег и Леша будут в разных группах
    - *Трое (многие-ко-многим):* `Аня К, Петя О, Маша И` → все трое в разных группах
    - *Один-ко-многим:* `Дмитрий К -> Елена В, Наталья С` → Дмитрий не должен быть с Еленой или Натальей, но Елена и Наталья могут быть вместе
    """)
    limits_text = st.text_area(
        "Введите ограничения",
        placeholder="""Примеры ограничений:
Олег С, Леша Ч                          ← многие-ко-многим: пара не должна быть вместе
Аня К, Петя О, Маша И                   ← многие-ко-многим: все трое в разных группах
Дмитрий К -> Елена В, Наталья С         ← один-ко-многим: Дмитрий отдельно от Елены и Натальи
Иван Д -> Ольга М                       ← один-ко-многим: Иван не должен быть с Ольгой""",
        height=240
    )
    
    # Settings
    settings = render_generation_settings()
    
    # Generate button
    st.markdown("---")
    run_button = st.button("🚀 РАСПРЕДЕЛИТЬ!", type="primary", width="stretch")
    
    if run_button:
        # Validate data
        df = editor_df
        if "Имя" not in df.columns or df.empty:
            st.error("Таблица пуста. Добавьте резидентов.")
            st.stop()
        
        valid_df = df.dropna(subset=["Имя"])
        names = valid_df["Имя"].str.strip().tolist()
        
        if len(set(names)) != len(names):
            st.error("В таблице есть дубликаты имён.")
            st.stop()
        
        # Extract data
        genders = dict(zip(valid_df["Имя"].str.strip(), valid_df["Пол"]))
        newbies = valid_df[valid_df["Роль"] == "Новичок"]["Имя"].str.strip().tolist()
        experts = valid_df[valid_df["Роль"] == "ВПИ"]["Имя"].str.strip().tolist()
        limits = parse_limits(limits_text)
        
        try:
            # Generate groups
            result = generate_groups(
                n=settings["num_groups"],
                all_people=names,
                genders=genders,
                newbies=newbies,
                experts=experts,
                limits=limits,
                seed=int(settings["seed"]) if settings["seed"] else None,
                strict_r=settings["strict_roles"],
                strict_g=settings["strict_genders"],
            )
            
            # Display results
            st.success(f"✅ Seed: {result.used_seed} | Попыток: {result.attempts}")
            if result.warnings:
                st.warning("⚠️ " + "; ".join(result.warnings))
            
            # Create formatted HTML for UI display and styled Excel export
            for i, group in enumerate(result.groups, 1):
                st.subheader(f"Группа {i} ({len(group)} чел.)")
                
                # Sort group members alphabetically
                sorted_group = sorted(group, key=lambda x: x.strip().lower())
                
                # Format names with styling for UI
                formatted_names = []
                for name in sorted_group:
                    name_stripped = name.strip()
                    is_vpi = name_stripped in experts
                    is_newbie = name_stripped in newbies
                    gender = genders.get(name_stripped, "M")
                    
                    # Build HTML with bold for VPI, italic for Newbies, and color for gender
                    # Female = red, Male = blue (UI only)
                    style_parts = []
                    if is_vpi:
                        style_parts.append("font-weight: bold;")
                    if is_newbie:
                        style_parts.append("font-style: italic;")
                    if gender == "F":
                        style_parts.append("color: red;")
                    elif gender == "M":
                        style_parts.append("color: blue;")
                    
                    if style_parts:
                        formatted_names.append(f'<span style="{" ".join(style_parts)}">{name_stripped}</span>')
                    else:
                        formatted_names.append(name_stripped)
                
                st.write(", ".join(formatted_names), unsafe_allow_html=True)
            
            # Export to Excel with professional formatting on a single sheet
            if result.groups:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Группы"
                
                # Build lookups from the current dataframe
                role_lookup = {}
                gender_lookup = {}
                current_df = st.session_state[DATA_KEY]
                for _, row in current_df.iterrows():
                    name = row["Имя"].strip() if isinstance(row["Имя"], str) else str(row["Имя"]).strip()
                    role_lookup[name] = row.get("Роль", "Обычный")
                    gender_lookup[name] = row.get("Пол", "M")
                
                # Define styles
                title_font = Font(size=16, bold=True)
                header_font = Font(size=12, bold=True)
                vpi_font = Font(bold=True)
                newbie_font = Font(italic=True)
                vpi_newbie_font = Font(bold=True, italic=True)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                center_alignment = Alignment(horizontal='center', vertical='center')
                left_alignment = Alignment(horizontal='left', vertical='center')
                
                # Calculate layout: each group gets 2 columns (title column + names column)
                # Group title spans both columns
                current_row = 1
                
                for group_idx, group in enumerate(result.groups, 1):
                    # Sort group members alphabetically
                    sorted_group = sorted(group, key=lambda x: x.strip().lower())
                    
                    # Column positions for this group (2 columns per group)
                    title_col = (group_idx - 1) * 2 + 1
                    names_col = title_col + 1
                    
                    # Set column widths
                    ws.column_dimensions[chr(64 + title_col)].width = 5
                    ws.column_dimensions[chr(64 + names_col)].width = 30
                    
                    # Add group title "N МАЛЫЕ ГРУППЫ" spanning both columns
                    title_text = f"{group_idx} МАЛЫЕ ГРУППЫ"
                    title_cell = ws.cell(row=current_row, column=title_col, value=title_text)
                    title_cell.font = title_font
                    title_cell.alignment = center_alignment
                    title_cell.border = thin_border
                    # Merge cells for title
                    ws.merge_cells(
                        start_row=current_row, 
                        end_row=current_row, 
                        start_column=title_col, 
                        end_column=names_col
                    )
                    
                    current_row += 1
                    
                    # Add column header "Группа N"
                    header_text = f"Группа {group_idx}"
                    header_cell = ws.cell(row=current_row, column=names_col, value=header_text)
                    header_cell.font = header_font
                    header_cell.alignment = left_alignment
                    header_cell.border = thin_border
                    
                    current_row += 1
                    
                    # Add sorted member names
                    for name in sorted_group:
                        name_cell = ws.cell(row=current_row, column=names_col, value=name)
                        name_cell.alignment = left_alignment
                        name_cell.border = thin_border
                        
                        # Apply font styling based on role
                        name_stripped = name.strip()
                        is_vpi = name_stripped in experts
                        is_newbie = name_stripped in newbies
                        
                        if is_vpi and is_newbie:
                            name_cell.font = vpi_newbie_font
                        elif is_vpi:
                            name_cell.font = vpi_font
                        elif is_newbie:
                            name_cell.font = newbie_font
                        
                        current_row += 1
                    
                    # Add empty row between groups for spacing (except after last group)
                    if group_idx < len(result.groups):
                        current_row += 1
                
                # Page setup for printing
                ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                ws.page_setup.paperSize = ws.PAPERSIZE_A4
                ws.page_margins.left = 0.5
                ws.page_margins.right = 0.5
                ws.page_margins.top = 0.75
                ws.page_margins.bottom = 0.75
                ws.print_options.horizontalCentered = True
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    "📥 Скачать результат в Excel",
                    data=output,
                    file_name="groups_result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        except Exception as e:
            st.error(f"❌ {e}")
        
        st.stop()


if __name__ == "__main__":
    main()
